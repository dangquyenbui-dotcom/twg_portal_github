"""
Excel Export Helper
Shared workbook builder for all sales report exports.
"""

from io import BytesIO
from datetime import date, datetime

from flask import session, send_file
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


def build_export_workbook(rows, title_label, columns, include_region_col=False):
    """
    Build a formatted Excel workbook from a list of row dicts.

    Args:
        rows:               List of dicts (one per row)
        title_label:        Report title string
        columns:            List of (header, dict_key, width, number_format) tuples
        include_region_col: If True, prepend a 'Region' column

    Returns:
        openpyxl Workbook
    """
    wb = Workbook()
    ws = wb.active
    ws.title = 'Raw Data'

    if include_region_col:
        cols = [('Region', 'Region', 10, None)] + list(columns)
    else:
        cols = list(columns)

    # ── Styles ──
    header_font = Font(name='Arial', bold=True, size=11, color='FFFFFF')
    header_fill = PatternFill('solid', fgColor='1F2937')
    header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell_font = Font(name='Arial', size=10)
    money_font = Font(name='Arial', size=10, color='0A7A4F')
    thin_border = Border(bottom=Side(style='thin', color='E5E7EB'))
    alt_fill = PatternFill('solid', fgColor='F9FAFB')

    MONEY_KEYS = ('ExtAmount', 'UnitPrice', 'OpenAmount', 'ExtPrice')

    # ── Title row ──
    today_str = date.today().strftime('%B %d, %Y')
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(cols))
    title_cell = ws.cell(row=1, column=1, value=f'{title_label} — {today_str}')
    title_cell.font = Font(name='Arial', bold=True, size=13, color='1F2937')
    title_cell.alignment = Alignment(vertical='center')
    ws.row_dimensions[1].height = 32

    # ── Exported by / timestamp ──
    user_name = session.get("user", {}).get("name", "Unknown")
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(cols))
    meta_cell = ws.cell(row=2, column=1,
                        value=f'Exported by {user_name} on {datetime.now().strftime("%m/%d/%Y %I:%M %p")}')
    meta_cell.font = Font(name='Arial', size=9, italic=True, color='6B7280')
    ws.row_dimensions[2].height = 20

    header_row = 4

    # ── Headers ──
    for col_idx, (header, _, width, _) in enumerate(cols, start=1):
        cell = ws.cell(row=header_row, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = Border(bottom=Side(style='medium', color='1F2937'))
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.row_dimensions[header_row].height = 28
    ws.freeze_panes = f'A{header_row + 1}'
    ws.auto_filter.ref = f'A{header_row}:{get_column_letter(len(cols))}{header_row + len(rows)}'

    # ── Data rows ──
    for row_idx, record in enumerate(rows, start=header_row + 1):
        is_alt = (row_idx - header_row) % 2 == 0
        for col_idx, (_, key, _, fmt) in enumerate(cols, start=1):
            value = record.get(key)
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = cell_font
            cell.border = thin_border
            cell.alignment = Alignment(vertical='center')

            if is_alt:
                cell.fill = alt_fill

            if fmt:
                cell.number_format = fmt

            if key in MONEY_KEYS:
                cell.font = money_font

    return wb


def send_workbook(wb, filename):
    """Save workbook to buffer and return as a downloadable response."""
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return send_file(
        buffer,
        as_attachment=True,
        download_name=filename,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.xml'
    )