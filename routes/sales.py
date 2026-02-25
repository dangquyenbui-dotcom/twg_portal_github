from io import BytesIO
from datetime import date, datetime

from flask import Blueprint, render_template, session, redirect, url_for, send_file
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from services.data_worker import get_bookings_from_cache
from services.db_service import fetch_bookings_raw, fetch_bookings_raw_ca

sales_bp = Blueprint('sales', __name__, url_prefix='/sales')


# ── Column config shared by all exports: (header, dict key, width, fmt) ──
EXPORT_COLUMNS = [
    ('Sales Order (sono)',        'SalesOrder',    20, None),
    ('Line# (tranlineno)',        'LineNo',        18, '#,##0'),
    ('Order Date (ordate)',       'OrderDate',     18, 'MM/DD/YYYY'),
    ('Customer No (custno)',      'CustomerNo',    20, None),
    ('Customer Name (company)',   'CustomerName',  30, None),
    ('Item (item)',               'Item',          18, None),
    ('Description (descrip)',     'Description',   32, None),
    ('Product Line (plinid)',     'ProductLine',   18, None),
    ('Qty Ordered (origqtyord)',  'QtyOrdered',    20, '#,##0'),
    ('Qty Shipped (qtyshp)',      'QtyShipped',    18, '#,##0'),
    ('Unit Price (price)',        'UnitPrice',     16, '$#,##0.00'),
    ('Ext Amount (calculated)',   'ExtAmount',     20, '$#,##0.00'),
    ('Ext Price (extprice)',      'ExtPrice',      18, '$#,##0.00'),
    ('Line Status (sostat)',      'LineStatus',    18, None),
    ('Order Type (sotype)',       'OrderType',     16, None),
    ('Territory (mapped)',        'Territory',     18, None),
    ('Terr Code (resolved)',      'TerrCode',      16, None),
    ('Tran Terr (tr.terr)',       'TranTerr',      16, None),
    ('SO Mast Terr (sm.terr)',    'SOMastTerr',    18, None),
    ('Cust Terr (cu.terr)',       'CustTerr',      16, None),
    ('Salesman (salesmn)',        'Salesman',      16, None),
    ('Location (loctid)',         'Location',      16, None),
    ('Request Date (rqdate)',     'RequestDate',   18, 'MM/DD/YYYY'),
    ('Ship Date (shipdate)',      'ShipDate',      18, 'MM/DD/YYYY'),
    ('Ship Via (shipvia)',        'ShipVia',       16, None),
]


def _build_export_workbook(rows, title_label, columns, include_region_col=False):
    """
    Build a formatted Excel workbook from a list of row dicts.
    Returns an openpyxl Workbook.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = 'Bookings Raw Data'

    # If combined export, prepend Region column
    if include_region_col:
        cols = [('Region', 'Region', 10, None)] + list(columns)
    else:
        cols = list(columns)

    # ── Styles ──
    header_font = Font(name='Arial', bold=True, size=11, color='FFFFFF')
    header_fill = PatternFill('solid', fgColor='1F2937')
    header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell_font = Font(name='Arial', size=10)
    money_font = Font(name='Arial', size=10, color='0A7A4F')
    thin_border = Border(bottom=Side(style='thin', color='E5E7EB'))
    alt_fill = PatternFill('solid', fgColor='F9FAFB')

    # ── Title row ──
    today_str = date.today().strftime('%B %d, %Y')
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(cols))
    title_cell = ws.cell(row=1, column=1, value=f'{title_label} — {today_str}')
    title_cell.font = Font(name='Arial', bold=True, size=13, color='1F2937')
    title_cell.alignment = Alignment(vertical='center')
    ws.row_dimensions[1].height = 32

    # ── Exported by / timestamp ──
    user_name = session.get("user", {}).get("name", "Unknown")
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(cols))
    meta_cell = ws.cell(row=2, column=1,
                        value=f'Exported by {user_name} on {datetime.now().strftime("%m/%d/%Y %I:%M %p")}')
    meta_cell.font = Font(name='Arial', size=9, italic=True, color='6B7280')
    ws.row_dimensions[2].height = 20

    header_row = 4

    # ── Headers ──
    for col_idx, (header, _, width, _) in enumerate(cols, start=1):
        cell = ws.cell(row=header_row, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = Border(bottom=Side(style='medium', color='1F2937'))
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.row_dimensions[header_row].height = 28
    ws.freeze_panes = f'A{header_row + 1}'
    ws.auto_filter.ref = f'A{header_row}:{get_column_letter(len(cols))}{header_row + len(rows)}'

    # ── Data rows ──
    for row_idx, record in enumerate(rows, start=header_row + 1):
        is_alt = (row_idx - header_row) % 2 == 0
        for col_idx, (_, key, _, fmt) in enumerate(cols, start=1):
            value = record.get(key)
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = cell_font
            cell.border = thin_border
            cell.alignment = Alignment(vertical='center')

            if is_alt:
                cell.fill = alt_fill

            if fmt:
                cell.number_format = fmt

            if key in ('ExtAmount', 'UnitPrice'):
                cell.font = money_font

    return wb


def _send_workbook(wb, filename):
    """Save workbook to buffer and return as a downloadable response."""
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return send_file(
        buffer,
        as_attachment=True,
        download_name=filename,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.xml'
    )


@sales_bp.route('/')
def sales_home():
    if not session.get("user"):
        return redirect(url_for('main.login_page'))
    return render_template('sales/index.html', user=session["user"])


@sales_bp.route('/bookings')
def bookings():
    if not session.get("user"):
        return redirect(url_for('main.login_page'))

    snapshot_us, snapshot_ca, last_updated = get_bookings_from_cache()

    # Build US data (or defaults)
    if snapshot_us is not None:
        us_summary = snapshot_us["summary"]
        us_data = {
            "total_amount": us_summary["total_amount"],
            "total_units": us_summary["total_units"],
            "total_orders": us_summary["total_orders"],
            "total_territories": us_summary["total_territories"],
            "territory_ranking": snapshot_us["ranking"],
            "order_date": us_summary.get("order_date"),
        }
    else:
        us_data = {
            "total_amount": 0, "total_units": 0, "total_orders": 0,
            "total_territories": 0, "territory_ranking": [],
            "order_date": None,
        }

    # Build CA data (or defaults)
    if snapshot_ca is not None:
        ca_summary = snapshot_ca["summary"]
        ca_data = {
            "total_amount": ca_summary["total_amount"],
            "total_units": ca_summary["total_units"],
            "total_orders": ca_summary["total_orders"],
            "total_territories": ca_summary["total_territories"],
            "territory_ranking": snapshot_ca["ranking"],
            "order_date": ca_summary.get("order_date"),
        }
    else:
        ca_data = {
            "total_amount": 0, "total_units": 0, "total_orders": 0,
            "total_territories": 0, "territory_ranking": [],
            "order_date": None,
        }

    error = None
    if snapshot_us is None and snapshot_ca is None:
        error = "Unable to load data. Please try again shortly."

    return render_template(
        'sales/bookings.html',
        user=session["user"],
        error=error,
        us=us_data,
        ca=ca_data,
        last_updated=last_updated
    )


@sales_bp.route('/bookings/export')
def bookings_export():
    """Export today's raw bookings data (US + Canada combined) as a formatted Excel file."""
    if not session.get("user"):
        return redirect(url_for('main.login_page'))

    rows_us = fetch_bookings_raw() or []
    rows_ca = fetch_bookings_raw_ca() or []

    if not rows_us and not rows_ca:
        return redirect(url_for('sales.bookings'))

    for row in rows_us:
        row['Region'] = 'US'
    for row in rows_ca:
        row['Region'] = 'CA'

    all_rows = rows_us + rows_ca

    wb = _build_export_workbook(
        rows=all_rows,
        title_label='Daily Bookings Raw Data (US + Canada)',
        columns=EXPORT_COLUMNS,
        include_region_col=True
    )

    filename = f'Bookings_Raw_US_CA_{date.today().strftime("%Y%m%d")}.xlsx'
    return _send_workbook(wb, filename)


@sales_bp.route('/bookings/export/us')
def bookings_export_us():
    """Export today's raw US bookings data as a formatted Excel file."""
    if not session.get("user"):
        return redirect(url_for('main.login_page'))

    rows = fetch_bookings_raw()
    if not rows:
        return redirect(url_for('sales.bookings'))

    wb = _build_export_workbook(
        rows=rows,
        title_label='Daily Bookings Raw Data — United States',
        columns=EXPORT_COLUMNS,
        include_region_col=False
    )

    filename = f'Bookings_Raw_US_{date.today().strftime("%Y%m%d")}.xlsx'
    return _send_workbook(wb, filename)


@sales_bp.route('/bookings/export/ca')
def bookings_export_ca():
    """Export today's raw Canada bookings data as a formatted Excel file."""
    if not session.get("user"):
        return redirect(url_for('main.login_page'))

    rows = fetch_bookings_raw_ca()
    if not rows:
        return redirect(url_for('sales.bookings'))

    wb = _build_export_workbook(
        rows=rows,
        title_label='Daily Bookings Raw Data — Canada',
        columns=EXPORT_COLUMNS,
        include_region_col=False
    )

    filename = f'Bookings_Raw_CA_{date.today().strftime("%Y%m%d")}.xlsx'
    return _send_workbook(wb, filename)